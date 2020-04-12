
#python strs.py --file1=c:\projects\string_similarity\data\data_1.txt --file2=c:\projects\string_similarity\data\data_2.txt --alg=0 --debug

"""************************************************************************************************
***************************************************************************************************
************************************************************************************************"""
import os
import sys
import string
import random
from pprint                    import pprint
from openpyxl                  import Workbook
from argparse                  import ArgumentParser
from strs_levenshtein          import STRS_ALG_Levenshtein
from strs_hamming              import STRS_ALG_Hamming
from strs_damerau_levenshtein  import STRS_ALG_Damerau_Levenshtein
from strs_gotoh                import STRS_ALG_Gotoh
from strs_jaro                 import STRS_ALG_Jaro
from strs_jaro_winkler         import STRS_ALG_Jaro_Winkler
from strs_mlipns               import STRS_ALG_Mlipns
from strs_needleman_wunsch     import STRS_ALG_Needleman_Wunsch
from strs_smith_waterman       import STRS_ALG_Smith_Waterman
from strs_strcmp95             import STRS_ALG_Strccmp95
from strs_tversky              import STRS_ALG_Tversky
from strs_tanimoto             import STRS_ALG_Tanimoto
from strs_overlap              import STRS_ALG_Overlap
from strs_monge_elkan          import STRS_ALG_Monge_Elkan
from strs_jaccard              import STRS_ALG_Jaccard
from strs_sorensen             import STRS_ALG_Sorenson
from strs_bag                  import STRS_ALG_Bag
from strs_cosine               import STRS_ALG_Cosine

"""************************************************************************************************
***************************************************************************************************
************************************************************************************************"""
_ALG_MAP = [
                STRS_ALG_Levenshtein,
                STRS_ALG_Hamming,
                STRS_ALG_Damerau_Levenshtein,
                STRS_ALG_Gotoh,
                STRS_ALG_Jaro,
                STRS_ALG_Jaro_Winkler,
                STRS_ALG_Mlipns,
                STRS_ALG_Needleman_Wunsch,
                STRS_ALG_Smith_Waterman,
                STRS_ALG_Strccmp95,
                STRS_ALG_Tversky,
                #STRS_ALG_Tanimoto,
                STRS_ALG_Overlap,
                STRS_ALG_Monge_Elkan,
                STRS_ALG_Jaccard,
                STRS_ALG_Sorenson,
                #STRS_ALG_Bag,
                STRS_ALG_Cosine
            ]

"""************************************************************************************************
***************************************************************************************************
************************************************************************************************"""
class STRS_Arguments(object):

    arguments = None

    def __init__(self):

        self.arg = ArgumentParser()

        self._add_arguments()

        self.arguments = self._parse_arguments()

    def _get_alg_help(self):

        _txt = ""

        for _key in range(len(_ALG_MAP)):

            _txt += "%s-%s  " % (_key,_ALG_MAP[_key].__name__.split("STRS_ALG_")[1].lower())

        return _txt

    def _add_arguments(self):

        self.arg.add_argument('--file1',       type=str,              help='first string file to compare')
        self.arg.add_argument('--file2',       type=str,              help='second string file to compare')  
        self.arg.add_argument('--alg',         type=str,              help=self._get_alg_help())
        self.arg.add_argument('--debug',       action='store_true',   help='verbose') 
        self.arg.add_argument('--stats',       action='store_true',   help='create stats for all algorithms')

    def _parse_arguments(self):

        if len(sys.argv) == 1:

            self.arg.print_help()

            sys.exit(1)

        return self.arg.parse_args()

"""************************************************************************************************
***************************************************************************************************
************************************************************************************************"""
class STRS(STRS_Arguments):

    def __init__(self):

        STRS_Arguments.__init__(self)

    def run(self):

        if self.arguments.alg != None:

            if self.arguments.alg != "all":

                self.arguments.alg = int(self.arguments.alg)

                _result = self.run_alg(
                                int(self.arguments.alg),
                                self.arguments.file1,
                                self.arguments.file2,
                                self.arguments.debug
                            )

                print(_result)

            else:
                for _index in range(len(_ALG_MAP)):

                    _result = self.run_alg(
                                            _index,
                                            self.arguments.file1,
                                            self.arguments.file2,
                                            self.arguments.debug
                                        )

                    print(_result)
        else:
            if self.arguments.stats != None:

                self.stats()
         
    def run_alg(self,index,file1,file2,debug):

        _result = None

        if index in range(len(_ALG_MAP)):

            _alg = _ALG_MAP[index](debug)

            if os.path.exists(str(file1)):

                _str1 = _alg.read_file(file1)

                if os.path.exists(str(file2)):

                    _str2 = _alg.read_file(file2)

                    _result = _alg.compare(_str1,_str2)
                
                else:
                    print("error: file2 does not exist %s" % (file2))
            else:
                print("error: file1 does not exist %s" % (file1))
        else:
            print("error: unknown algorithm")

        return _result

    def get_data_pairs(self):

        _pairs = [
                    [r"..\data\data_1_0.txt",r"..\data\data_1_1.txt"],
                    [r"..\data\data_1_0.txt",r"..\data\data_1_2.txt"],
                    [r"..\data\data_1_0.txt",r"..\data\data_1_3.txt"],
                ]

        return _pairs

    def stats(self):

        _pairs = self.get_data_pairs()

        _workbook = Workbook()

        del _workbook[_workbook.active.title]

        #self.stats_table(woorkbook)

        self.stats_change_rate(woorkbook)

        _workbook.save("stats.xlsx")


    def stats_table(self,woorkbook)

        _sheet = woorkbook.create_sheet(title="stats")

        #add x axis
        _column = 2
        for _pair in _pairs:

            _file1 = os.path.splitext(os.path.split(_pair[0])[-1])[0]

            _file2 = os.path.splitext(os.path.split(_pair[1])[-1])[0]

            _sheet.cell(row=1,column=_column + 0).value = "%s - %s (distance)" % (_file1,_file2)

            _sheet.cell(row=1,column=_column + 1).value = "%s - %s (similarity)" % (_file1,_file2)

            _sheet.cell(row=1,column=_column + 2).value = "%s - %s (similar)" % (_file1,_file2)

            _sheet.cell(row=1,column=_column + 3).value = "%s - %s (time)" % (_file1,_file2)

            for _index in range(len(_ALG_MAP)):

                _result = self.run_alg(
                                            _index,
                                            _pair[0],
                                            _pair[1],
                                            self.arguments.debug
                                        )

                _sheet.cell(row=1 + (_index + 1), column=_column + 0).value = _result.distance
                _sheet.cell(row=1 + (_index + 1), column=_column + 1).value = _result.similarity
                _sheet.cell(row=1 + (_index + 1), column=_column + 2).value = _result.similar
                _sheet.cell(row=1 + (_index + 1), column=_column + 3).value = _result.time

            _column += 4

        #add y axis
        _row = 2
        for _key in range(len(_ALG_MAP)):
            _sheet.cell(row=_row,column=1).value = "%s" % (_ALG_MAP[_key].__name__.split("STRS_ALG_")[1].lower(),)
            _row += 1

    def stats_change_rate(self,woorkbook):

        _file = r"..\data\data_1_0.txt"


        _new_ch = random.choice(string.ascii_letters)

        print(_new_ch)


"""************************************************************************************************
***************************************************************************************************
************************************************************************************************"""
if __name__ == '__main__':

    _strs = STRS()

    _strs.run()